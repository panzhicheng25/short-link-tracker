from flask import Blueprint, request, jsonify, Response
from models import (create_link, get_all_links_with_last_click, get_link_by_id,
                    delete_link, batch_create_links, export_clicks_by_link, batch_generate_links)
from utils.short_code import encode
import time
from io import BytesIO
from openpyxl import Workbook, load_workbook

links_bp = Blueprint('links', __name__)

# ========== Excel 导入（仅解析，不写DB） ==========

@links_bp.route('/api/links/import-excel', methods=['POST'])
def import_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "no file"}), 400
    try:
        wb = load_workbook(file)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
            title = str(row[0]).strip() if row[0] else ''
            url = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if url:
                rows.append({"title": title, "original_url": url})
        wb.close()
        return jsonify({"rows": rows, "count": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ========== 批量生成短链（带剧名） ==========

@links_bp.route('/api/links/batch-generate', methods=['POST'])
def batch_generate():
    data = request.get_json() or {}
    links_data = data.get('links', [])
    if not links_data:
        return jsonify({"error": "links required"}), 400
    results = batch_generate_links(links_data)
    return jsonify({"results": results, "count": len(results)})

# ========== 单条创建 ==========

@links_bp.route('/api/links', methods=['POST'])
def create_new_link():
    data = request.get_json() or {}
    original_url = data.get('original_url', '').strip()
    title = data.get('title', '').strip()
    if not original_url:
        return jsonify({"error": "original_url required"}), 400
    temp_id = int(time.time() * 1000) % 1000000000
    short_code = encode(temp_id)
    try:
        link_id = create_link(short_code, original_url, title)
    except:
        short_code = encode(temp_id + 999999)
        link_id = create_link(short_code, original_url, title)
    return jsonify({"id": link_id, "short_code": short_code, "title": title, "original_url": original_url})

# ========== 旧的批量导入（纯链接） ==========

@links_bp.route('/api/links/batch', methods=['POST'])
def batch_create():
    data = request.get_json() or {}
    text = data.get('text', '').strip()
    if not text:
        return jsonify({"error": "text required"}), 400
    urls = [line.strip() for line in text.split('\n') if line.strip()]
    results = batch_create_links(urls)
    return jsonify({"results": results, "count": len(results)})

# ========== 下载全部短链（三列：剧名、分销链接、短链） ==========

@links_bp.route('/api/links/download-all')
def download_all():
    links = get_all_links_with_last_click()
    wb = Workbook()
    ws = wb.active
    ws.title = "全部短链"
    from config import BASE_URL
    ws.append(['剧名', '分销链接', '短链'])
    for l in links:
        full_url = f"{BASE_URL}/s/{l.get('short_code', '')}"
        ws.append([l.get('title', ''), l.get('original_url', ''), full_url])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=all_links.xlsx'})

# ========== 列表 ==========

@links_bp.route('/api/links', methods=['GET'])
def list_links():
    return jsonify(get_all_links_with_last_click())

@links_bp.route('/api/links/<int:link_id>', methods=['GET'])
def get_link(link_id):
    link = get_link_by_id(link_id)
    if not link:
        return jsonify({"error": "Link not found"}), 404
    return jsonify(link)

@links_bp.route('/api/links/<int:link_id>', methods=['DELETE'])
def remove_link(link_id):
    delete_link(link_id)
    return jsonify({"success": True})


@links_bp.route('/api/links/export-selected', methods=['POST'])
def export_selected():
    """导出选中的多条短链数据为 Excel"""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({"error": "ids required"}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "选中短链"
    ws.append(['剧名', '分销链接', '短链', '点击量', '创建时间'])
    
    from config import BASE_URL as URL
    for lid in ids:
        link = get_link_by_id(lid)
        if link:
            full_url = f"{URL}/s/{link.get('short_code','')}"
            ws.append([link.get('title',''), link.get('original_url',''),
                       full_url, link.get('total_clicks',0),
                       link.get('created_at','')])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=selected_links.xlsx'})

# ========== 导出单条短链（剧名） ==========

@links_bp.route('/api/links/<int:link_id>/export')
def export_link(link_id):
    data = export_clicks_by_link(link_id)
    link = data['link']
    clicks = data['clicks']
    wb = Workbook()
    ws = wb.active
    ws.title = "点击数据"
    ws.append(['短链', '剧名', '分销链接'])
    ws.append([link.get('short_code',''), link.get('title',''), link.get('original_url','')])
    ws.append([])
    ws.append(['点击时间', '国家', '省份', '城市', '设备', '浏览器', '操作系统', '来源'])
    for c in clicks:
        ws.append([c.get('clicked_at',''), c.get('country',''), c.get('region',''),
                   c.get('city',''), c.get('device',''), c.get('browser',''),
                   c.get('os',''), c.get('referrer','')])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=link_{link_id}_export.xlsx'})
